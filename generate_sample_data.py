#!/usr/bin/env python3
"""Generate sample-data.xlsx for Cowork onboarding Card 4.

3 tabs of 12-month fake data with clear trends:
- Sales: growth trend with seasonal Q4 bump
- Marketing: one channel clearly outperforming, one underperforming
- Support: ticket volume rising, resolution time improving, one escalation spike
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample-data.xlsx")

HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
DISCLAIMER_FONT = Font(bold=True, size=11, color="C62828")
DISCLAIMER_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

MONTHS = [
    "Jan 2025", "Feb 2025", "Mar 2025", "Apr 2025", "May 2025", "Jun 2025",
    "Jul 2025", "Aug 2025", "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025",
]


def style_header_row(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def add_disclaimer(ws, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1)
    cell.value = "SAMPLE DATA - NOT REAL"
    cell.font = DISCLAIMER_FONT
    cell.fill = DISCLAIMER_FILL
    cell.alignment = Alignment(horizontal="center")


def create_sales_tab(wb):
    ws = wb.active
    ws.title = "Sales"
    headers = ["Month", "Region", "Units Shipped", "Revenue", "Avg Order Value", "Fulfillment Days", "Return Rate"]
    add_disclaimer(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, len(headers))

    regions = ["West", "Southeast", "Northeast", "Midwest"]
    base_units = [42, 38, 45, 40, 48, 44, 50, 52, 55, 58, 72, 85]
    base_aov = [4200, 4250, 4180, 4300, 4350, 4280, 4400, 4320, 4450, 4500, 4380, 4550]
    fulfill = [5.2, 4.8, 4.5, 4.3, 4.1, 3.9, 3.8, 3.7, 3.5, 3.4, 3.6, 4.0]
    returns = [0.032, 0.028, 0.025, 0.024, 0.022, 0.020, 0.019, 0.018, 0.017, 0.016, 0.018, 0.015]

    row = 3
    for i, month in enumerate(MONTHS):
        for region in regions:
            multiplier = {"West": 1.0, "Southeast": 0.85, "Northeast": 0.75, "Midwest": 0.65}[region]
            units = int(base_units[i] * multiplier)
            aov = int(base_aov[i] * (1 + (0.05 if region == "West" else 0)))
            revenue = units * aov
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=region)
            ws.cell(row=row, column=3, value=units)
            ws.cell(row=row, column=4, value=revenue)
            ws.cell(row=row, column=5, value=aov)
            ws.cell(row=row, column=6, value=round(fulfill[i] + (0.3 if region == "Midwest" else 0), 1))
            ws.cell(row=row, column=7, value=round(returns[i] * (1.2 if region == "Midwest" else 1.0), 3))
            row += 1


    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def create_marketing_tab(wb):
    ws = wb.create_sheet("Marketing")
    headers = ["Month", "Channel", "Spend", "Impressions", "Clicks", "Conversions", "CAC", "Website Sessions"]
    add_disclaimer(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, len(headers))

    channels = {
        "Google Ads": {"spend": 18000, "imp": 450000, "ctr": 0.035, "cvr": 0.028, "sessions": 15750, "trend": 1.08},
        "Facebook": {"spend": 12000, "imp": 380000, "ctr": 0.018, "cvr": 0.012, "sessions": 6840, "trend": 0.95},
        "Instagram": {"spend": 8000, "imp": 220000, "ctr": 0.022, "cvr": 0.015, "sessions": 4840, "trend": 1.12},
        "SEO/Organic": {"spend": 3500, "imp": 180000, "ctr": 0.042, "cvr": 0.035, "sessions": 7560, "trend": 1.15},
        "Email": {"spend": 1500, "imp": 45000, "ctr": 0.065, "cvr": 0.048, "sessions": 2925, "trend": 1.05},
    }

    row = 3
    for i, month in enumerate(MONTHS):
        for name, ch in channels.items():
            growth = ch["trend"] ** i
            spend = int(ch["spend"] * growth)
            impressions = int(ch["imp"] * growth)
            clicks = int(impressions * ch["ctr"])
            conversions = int(clicks * ch["cvr"])
            cac = round(spend / max(conversions, 1))
            sessions = int(ch["sessions"] * growth)
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=spend)
            ws.cell(row=row, column=4, value=impressions)
            ws.cell(row=row, column=5, value=clicks)
            ws.cell(row=row, column=6, value=conversions)
            ws.cell(row=row, column=7, value=cac)
            ws.cell(row=row, column=8, value=sessions)
            row += 1


    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


def create_support_tab(wb):
    ws = wb.create_sheet("Support")
    headers = ["Month", "Tickets", "Avg Resolution Hours", "CSAT Score", "Top Category", "Escalation Rate"]
    add_disclaimer(ws, len(headers))
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header_row(ws, len(headers))

    tickets = [145, 152, 168, 175, 182, 190, 198, 210, 225, 238, 260, 285]
    resolution = [18.5, 17.2, 16.8, 15.5, 14.8, 14.2, 13.5, 12.8, 12.2, 11.5, 11.8, 10.5]
    csat = [4.1, 4.15, 4.2, 4.25, 4.3, 4.35, 4.4, 4.45, 4.5, 4.55, 4.5, 4.6]
    categories = [
        "Setup & Installation", "Setup & Installation", "Billing/Pricing",
        "Billing/Pricing", "Billing/Pricing", "Product Questions",
        "Product Questions", "Setup & Installation", "Billing/Pricing",
        "Billing/Pricing", "Setup & Installation", "Setup & Installation",
    ]
    escalation = [0.12, 0.11, 0.10, 0.09, 0.085, 0.08, 0.075, 0.07, 0.065, 0.18, 0.07, 0.06]

    row = 3
    for i, month in enumerate(MONTHS):
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=tickets[i])
        ws.cell(row=row, column=3, value=resolution[i])
        ws.cell(row=row, column=4, value=csat[i])
        ws.cell(row=row, column=5, value=categories[i])
        ws.cell(row=row, column=6, value=escalation[i])
        row += 1


    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def generate():
    wb = Workbook()
    create_sales_tab(wb)
    create_marketing_tab(wb)
    create_support_tab(wb)
    wb.save(OUTPUT)
    print(f"Done - wrote {OUTPUT}")


if __name__ == "__main__":
    generate()
